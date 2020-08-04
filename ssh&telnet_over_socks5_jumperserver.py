import paramiko
import socks
import time
import socket
import sys
import telnetlib
from openpyxl import Workbook

from openpyxl import load_workbook

from openpyxl.writer.excel import ExcelWriter

socks.setdefaultproxy(socks.PROXY_TYPE_SOCKS5, 'proxyip', proxyport, False, 'username', 'passwd') # socks5 jumper server
socket.socket = socks.socksocket


ssh = paramiko.SSHClient() # instance ssh


ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())

WB = load_workbook('excel file path')
i = 0
WS = WB['IP']

wbA = Workbook()
wsA = wbA.active
while 1 == 1:
    i = i + 1
    hostname = WS.cell(i, 1).value


    if not hostname == None:


        try:
            ssh.connect(hostname, port=22, username='XXXX', password='XXXX', timeout=5)
            remote_conn = ssh.invoke_shell()
            remote_conn.send('enable\n')
            time.sleep(1)
            remote_conn.send('password for enable\n')

            time.sleep(1)
            # remote_conn.send('show run | section router\n')
            # time.sleep(15)
            remote_conn.send('terminal length 0\n')
            time.sleep(1)
            remote_conn.send('show ip int | in line protocol|Internet address\n')
            # remote_conn.send('show vlan\n')
            time.sleep(15)
            routeospf = remote_conn.recv(65535)
            print(routeospf.decode())
            remote_conn.send('show run | in ip route\n')
            # remote_conn.send('show int trunk\n')
            time.sleep(15)
            iproute = remote_conn.recv(65535)
            print(iproute.decode())
            # print(result.decode())
            # result = ''
            # remote_conn.send('show run | in mask\n')
            # remote_conn.send('show interface status\n')
            # print('11111')
            # time.sleep(10) # time need to modify depends on switch or router
            # routebgp = remote_conn.recv(65535)
            # print(result)
            # print(routebgp.decode())
            # if not (routeospf or iproute or routebgp):
            if not (routeospf or iproute):
                route = sys.stderr.read()
                ssh.close()
            print(hostname)
            # print(routeospf.decode())
            # print(iproute.decode())
            # print(routebgp.decode())

            wsA.cell(i, 1).value = hostname
            print(wsA.cell(i, 1).value)
            print(type(wsA.cell(i, 1).value))
            wsA.cell(i, 2).value = routeospf.decode()
            wsA.cell(i, 3).value = iproute.decode()

            wbA.save('excel of result file path')
            ssh.close()
        except (paramiko.ssh_exception.AuthenticationException):
            print('Authentication failed')
        except (socks.GeneralProxyError):  # try telnet once ssh failed
            tn = telnetlib.Telnet(hostname, port=23, timeout=5)
            tn.read_until(('sername:').encode())  # match the promote

            tn.write(('username for telnet\n').encode())  # encoding the string
            tn.read_until(('assword:').encode()) # match the promote

            tn.write(('password for telnet\n').encode())
            tn.write(('enable\n').encode())

            tn.write(('enable password for machine\n').encode())
            tn.write(('terminal length 0\n').encode())
            time.sleep(1)


            tn.write(('show ip int | in line protocol|Internet address\n').encode())
            time.sleep(15)
            routeospf = tn.read_very_eager().decode()
            print(routeospf)
            # print(result.decode())
            # result = ''
            tn.write(('show run | in ip route\n').encode())
            time.sleep(10)
            iproute = tn.read_very_eager().decode()
            print(iproute)

            if not (routeospf or iproute):
                # result = stderr.read()
                tn.close()
            print(hostname)

            wsA.cell(i, 1).value = hostname
            wsA.cell(i, 2).value = routeospf
            wsA.cell(i, 3).value = iproute
            wbA.save('C:\\Users\......excel of result file path')
            tn.close()
        except:
            print(hostname, 'not good')
            wsA.cell(i, 1).value = hostname
            wsA.cell(i, 2).value = ('not good')
            wbA.save('C:\\Users\......excel of result file path')
            continue


    else:
        break
